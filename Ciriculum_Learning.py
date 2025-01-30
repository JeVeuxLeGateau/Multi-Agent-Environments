import ray
from ray.rllib.algorithms.ppo import PPOConfig
from ray.tune.registry import register_env
from ray.rllib.models import ModelCatalog
import gym_FTP as e
import pygame
import os
import matplotlib.pyplot as plt
from CustomCallback import CustomCallback
import Train_Critic as tc
import Define_critic as dc
import torch
from openpyxl import Workbook
import pandas as pd
from ray.rllib.algorithms.algorithm import Algorithm
import json
import misc as m
import numpy as np
from sklearn.linear_model import LogisticRegression
import statistics

os.environ["PYTHONWARNINGS"] = "ignore::DeprecationWarning"


# Environment creation function
def env_creator(config):
    robots = config.get("robots", [])  # Provide a default value if not found
    adversaries = config.get("adversaries", [])

    if not pygame.get_init():
        pygame.init()
    screen = pygame.display.set_mode([1000, 1000])

    time_steps = 500
    overseer = e.Overseer(screen, robots, 0, adversaries, time_steps, 15)
    return overseer


def toggle_eval(file_name, new_value1, new_value2):
    if os.path.exists(file_name):
        df_existing = pd.read_excel(file_name)
        df_existing.loc[0, 'eval_true'] = new_value1
        df_existing.loc[0, 'render'] = new_value2
        df_existing.to_excel(file_name, index=False)

    else:
        print(f"File {file_name} does not exist.")


def train_and_evaluate(learning_episodes, time_steps, eval_episodes, robots, adversaries, learning_rate, gamma, ec, cp,
                       tbs, mbs):
    ray.init(ignore_reinit_error=True)
    register_env("Env_FTP", lambda config: env_creator(config={"robots": robots, "adversaries": adversaries}))
    ModelCatalog.register_custom_model("custom_critic", dc.CustomCritic)

    config = (PPOConfig()
              .environment("Env_FTP", env_config={"robots": robots, "adversaries": adversaries, "is_evaluation": False})
              .env_runners(
        num_env_runners=0,  # Use num_env_runners instead of num_rollout_workers
        rollout_fragment_length=time_steps,
        create_env_on_local_worker=True)
              .training(
        train_batch_size=tbs,
        minibatch_size=mbs,
        num_epochs=15,
        model={
            "custom_model": "custom_critic",
            "vf_share_layers": False,

            "custom_model_config": {
                "agents": robots + adversaries,  # Total number of agents
                "robots": robots,
                "adversaries": adversaries,
                "pretrained_critic_path": "pretrained_critic.pth",  # Path to your saved critic
                "best_model_params_path": "best_model_params_path.pth"
            },

        },
        lr=learning_rate,  # Learning rate
        gamma=gamma,
        entropy_coeff=ec,
        clip_param=cp
    ).reporting(keep_per_episode_custom_metrics=True)
              .api_stack(  # Use api_stack
        enable_rl_module_and_learner=False
    )
              .framework("torch")
              .evaluation(evaluation_num_env_runners=1,
                          evaluation_duration=eval_episodes,
                          evaluation_config={"env_config": {"is_evaluation": True}})
              .resources(num_gpus=0)
              .callbacks(CustomCallback))

    """
    algo = config.build()
    print("End Build")
    """

    # Refresh workbooks

    wb1 = Workbook()
    wb1.save("Results_from_Training.xlsx")

    wb2 = Workbook()
    wb2.save("Successes.xlsx")

    wb3 = Workbook()
    wb3.save("metric_values.xlsx")

    data = {'eval_true': ["False"],
            'render': ["True"]}  # Start with True

    df = pd.DataFrame(data)
    file_name = "eval_file.xlsx"
    df.to_excel(file_name, index=False)

    """
    current_file = "4_0_best_weights.pth"

    best_reward = float('-inf')
    for episode in range(learning_episodes):
        print("Episode:", episode)
        results = algo.train()
        current_reward = results["env_runners"]["episode_reward_mean"]

        if best_reward < current_reward:
            policy = algo.get_policy()
            weights = {k: v for k, v in policy.model.state_dict().items()}
            best_reward = current_reward
            print("New Best Weights")

            torch.save(weights, current_file)
        print("===============================================================")

    print("End of training")
    print("Saving end weights")

    current_file = "4_0_end_weights.pth"

    policy = algo.get_policy()
    weights = {k: v for k, v in policy.model.state_dict().items()}
    torch.save(weights, current_file)

    episodes_list = list(range(1, learning_episodes + 1))

    m.plot_results(episodes_list)

    print()
    names = ["wins/loss"]
    df = pd.read_excel("Successes.xlsx", sheet_name="Sheet", names=names)
    wins = (df["wins/loss"] == 1).sum()

    print("SUCCESS RATE FROM TRAINING:", wins / (learning_episodes * 5))
    print()

    toggle_eval(file_name, "True", "True")
    evaluation_results = algo.evaluate()

    print("EVALUATION RESULTS:",
          evaluation_results["env_runners"]["custom_metrics"]["last_termination_value"].count(1) / eval_episodes)

    print()
    print("Now with best")
    print()
    """
    file_name = "eval_file.xlsx"
    toggle_eval(file_name, "True", "True")
    algo2 = config.build()
    evaluation_results = algo2.evaluate()

    print("EVALUATION RESULTS WITH BEST:",
          evaluation_results["env_runners"]["custom_metrics"]["last_termination_value"].count(1) / eval_episodes)

    df = pd.read_excel("metric_values.xlsx")

    # Convert rows to a list of lists
    # 0 - win/loss
    # 1 - # of adversaries defeated
    # 2 - # of robots unfrozen
    # 3 - Final time
    rows = df.values.tolist()

    # Now iterate through the rows
    survived = 0
    survived_list = []
    adversaries_defeated = 0
    defeated_list = []
    robots_unfrozen = 0
    time_steps = 0
    time_step_list = []
    timeouts = 0
    frozen_loss = 0
    ds_list = []
    MDS_time = []

    for row in rows:

        if row[0] == 1:
            survived += 1
            survived_list.append(1)
            time_steps += row[3]
            print("ds", row[5])
            print("time", (row[3] * 10) + row[4])
            print("Ratio:", ((row[3] * 10) + row[4]) / row[5])
            MDS_time.append((row[3] * 10) + row[4])
            ds_list.append(row[5])
            print()
        elif row[0] == -1:
            timeouts += 1
            survived_list.append(0)
        elif row[0] == -2:
            frozen_loss += 1
            survived_list.append(0)

        adversaries_defeated += row[1]
        defeated_list.append(row[1])
        robots_unfrozen += row[2]
        time_step_list.append((row[3] * 10) + row[4])

    print()
    print("Robots", robots)
    print("Adversaries:", adversaries)
    print("Total Episodes:", eval_episodes)
    print("Survival Rate:", survived / eval_episodes)
    print("Timeout Rate:", timeouts / eval_episodes)
    print("All Frozen Loss:", frozen_loss / eval_episodes)
    print("Adversary Defeat Rate:", adversaries_defeated / eval_episodes)
    print("Robot Unfreeze Rate:", robots_unfrozen / eval_episodes)
    print("Average Time steps:", time_steps / eval_episodes)

    X1 = np.array(defeated_list).reshape(-1, 1)
    y1 = np.array(survived_list)
    X2 = np.array(time_step_list).reshape(-1, 1)
    y2 = np.array(survived_list)

    B_0_a = "None"
    B_1_a = "None"
    B_0_t = "None"
    B_1_t = "None"

    if len(set(y1)) == 2:
        model = LogisticRegression()
        model.fit(X1, y1)
        B_0_a = model.intercept_[0]
        B_1_a = model.coef_[0][0]
        print("B_0_a:", B_0_a)
        print("B_1_a:", B_1_a)

    if len(set(y2)) == 2:
        model = LogisticRegression()
        model.fit(X2, y2)
        B_0_t = model.intercept_[0]
        B_1_t = model.coef_[0][0]
        print("B_0_t:", B_0_t)
        print("B_1_t:", B_1_t)

    MDS_ratios = []

    MDS = 0
    for i in range(len(MDS_time)):
        MDS += MDS_time[i] / ds_list[i]
        MDS_ratios.append(MDS_time[i] / ds_list[i])

    MDS = MDS / len(time_step_list)
    print("MDS:", MDS)
    print("median:", statistics.median(MDS_ratios))
    print("min_time:", min(MDS_ratios))
    print("max_time:", max(MDS_ratios))

    # Shutdown Ray
    ray.shutdown()


def main():
    # num_critic_training_episodes = 1000
    num_PPO_training_episodes = 1500
    num_time_steps = 500
    num_evaluate_episodes = 1000
    robots = 2
    adversaries = 1
    alpha = 0.0001
    gamma = 0.95
    entropy = 0.015
    clip_param = 0.2
    training_batch_size = 2500
    minibatch_size = 500

    """
    # Train critic
    print("Training the critic...")


    trained_critic = tc.train_critic_loop(robots, adversaries, num_critic_training_episodes, num_time_steps, gamma,
                                          alpha)
    weights = trained_critic.state_dict()
    critic_weights = {k: v for k, v in weights.items() if k.startswith("critic")}
    modified_weights = {k.replace('critic.', ''): v for k, v in critic_weights.items()}

    # Save Critic
    torch.save(modified_weights, "pretrained_critic.pth")
    """

    train_and_evaluate(num_PPO_training_episodes, num_time_steps, num_evaluate_episodes, robots, adversaries,
                       alpha, gamma, entropy, clip_param,
                       training_batch_size, minibatch_size)


main()

"""
Best Results:
=============
2_0_end_weights.pth
3_0_end_weights_retrained4.pth
4_0.pth
2_1_end_weights.pth
3_1.pth

"""
