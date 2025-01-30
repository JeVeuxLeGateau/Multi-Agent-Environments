from ray.rllib.algorithms.callbacks import DefaultCallbacks
import openpyxl
from openpyxl import Workbook
import misc as m
import os
import pandas as pd


def add_data_to_first_empty_row(filename, data, sheet_name="Sheet1"):
    # Load the existing workbook
    wb = openpyxl.load_workbook(filename)

    # Access the specified sheet (or default to the active one if it doesn't exist)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # Find the first empty row
    empty_row = ws.max_row + 1  # Assuming no completely blank rows before this one

    # Write data to the first empty row
    for col, value in enumerate(data, start=1):
        ws.cell(row=empty_row, column=col, value=value)

    # Save the workbook with the new data
    wb.save(filename)


def print_dict(d, indent=0):
    for key, value in d.items():
        if isinstance(value, dict):
            print("\t" * indent + str(key) + ":")
            print_dict(value, indent + 1)
        else:
            print("\t" * indent + f"{key}: {value}")


class CustomCallback(DefaultCallbacks):

    def on_episode_start(self, *, worker, base_env, policies, episode, env_index, **kwargs):
        # Initialize storage for termination information
        episode.user_data["termination_vals"] = []
        n_C = worker.env.n_C
        n_E = worker.env.n_E

        episode.custom_metrics["n_C"] = n_C
        episode.custom_metrics["n_E"] = n_E

    def on_episode_step(self, *, worker, base_env, policies, episode, env_index, **kwargs):
        # Access info from kwargs
        info = episode._last_infos["agent0"]["val"]
        episode.user_data["termination_vals"].append(info)

    def on_episode_end(self, *, worker, base_env, policies, episode, env_index, **kwargs):
        # Analyze the termination_vals to count specific termination cases

        last_termination_value = episode.user_data["termination_vals"][-1]
        info = episode._last_infos["agent0"]
        CM = episode.custom_metrics
        metric_values = m.process_info(info, CM["n_C"], CM["n_E"])

        episode.custom_metrics["last_termination_value"] = last_termination_value
        episode.custom_metrics["metric_values"] = metric_values

        file_path = "metric_values.xlsx"

        # Convert data to a DataFrame
        data = {
            "Metric 1": [metric_values[0]],
            "Metric 2": [metric_values[1]],
            "Metric 3": [metric_values[2]],
            "Metric 4": [metric_values[3]],
            "offset": [info["offset"]],
            "dist": [info["dist"]]

        }
        df = pd.DataFrame(data)

        if os.path.exists(file_path):
            existing_df = pd.read_excel(file_path)
            updated_df = pd.concat([existing_df, df], ignore_index=True)
            updated_df.to_excel(file_path, index=False)
        else:
            df.to_excel(file_path, index=False)

        data = [last_termination_value]
        add_data_to_first_empty_row("Successes.xlsx", data)

    def on_train_result(self, *, algorithm, result, **kwargs):
        val = result['info']['learner']["default_policy"]["learner_stats"]
        episode_reward_mean = result["env_runners"]["episode_reward_mean"]
        total_loss = val["total_loss"]
        policy_loss = val["policy_loss"]
        value_function_loss = val["vf_loss"]
        kl = val["kl"]
        entropy = val["entropy"]
        entropy_coefficient = algorithm.config["entropy_coeff"]

        data = [episode_reward_mean, total_loss, policy_loss, value_function_loss, kl, entropy]
        add_data_to_first_empty_row("Results_from_Training.xlsx", data)

        print("Episode Reward Mean:", episode_reward_mean)
        print("Total Loss:", total_loss)
        print("Policy Loss:", policy_loss)
        print("Value Function Loss:", value_function_loss)
        print("KL:", kl)
        print("Entropy:", entropy)
        """
        print("Initial Entropy Coefficient:", entropy_coefficient)

        if entropy > 20:
            algorithm.config["entropy_coeff"] = max(entropy_coefficient * 0.9, 0.0001)  # Reduce coefficient
        elif entropy < 8:
            algorithm.config["entropy_coeff"] = min(entropy_coefficient * 1.1, 0.1)  # Increase coefficient
        print("Adjusted Entropy Coefficient:", algorithm.config["entropy_coeff"])
        """

